import os
import json
import argparse
import pandas as pd

"""
该脚本是针对使用第三方库 https://github.com/aissat/easy_localization 进行国际化的 Flutter 项目
"""

# Flutter 项目中 国际化文案 文件所在的目录，相对于脚本执行位置
DEFAULT_TRANSLATIONS_DIR = 'assets/translations'
# 默认的模板语言文件名
DEFAULT_TEMPLATE_LANG_FILE = 'en.json'


def export_translations(translations_dir, output_file):
    print(f"Exporting translations from {translations_dir} to {output_file}...")
    all_translations = {}
    languages = []
    master_keys = []

    # 确保 translations_dir 存在
    if not os.path.isdir(translations_dir):
        print(f"Error: Directory '{translations_dir}' not found.")
        return

    # 1. 读取所有 .json 文件并提取语言代码
    translations_files = [f for f in os.listdir(translations_dir) if f.endswith('.json')]

    if not translations_files:
        print(f"No .json files found in '{translations_dir}'.")
        return

    # 确定模板语言文件 (通常是 'en.json')
    template_file_name = DEFAULT_TEMPLATE_LANG_FILE
    if template_file_name not in translations_files:
        # 如果默认模板文件不存在，则选择列表中的第一个作为模板
        if translations_files:
            template_file_name = translations_files[0]
            print(
                f"Warning: Default template file '{DEFAULT_TEMPLATE_LANG_FILE}' not found. Using '{template_file_name}' as template.")
        else:
            print("Error: No .json files found to use as a template.")
            return

    template_lang_code = template_file_name.replace('.json', '')
    languages.append(template_lang_code)  # 模板语言放第一位

    # 读取模板文件以获取所有 key
    try:
        with open(os.path.join(translations_dir, template_file_name), 'r', encoding='utf-8') as f:
            template_data = json.load(f)
            master_keys = list(template_data.keys())
            all_translations[template_lang_code] = template_data
    except Exception as e:
        print(f"Error reading template file {template_file_name}: {e}")
        return

    # 读取其他语言文件
    for translations_file in translations_files:
        lang_code = translations_file.replace('.json', '')
        if lang_code == template_lang_code:  # 跳过已处理的模板文件
            continue
        languages.append(lang_code)
        try:
            with open(os.path.join(translations_dir, translations_file), 'r', encoding='utf-8') as f:
                all_translations[lang_code] = json.load(f)
        except Exception as e:
            print(f"Error reading file {translations_file}: {e}")
            # 即使某个文件读取失败，也继续处理其他文件，但记录错误
            all_translations[lang_code] = {}

    # 2. 创建 DataFrame
    df_data = {'key': master_keys}

    # 3. 填充翻译数据
    for lang in languages:
        lang_translations = all_translations.get(lang, {})
        column_data = []
        for key in master_keys:
            value = lang_translations.get(key, '')
            if isinstance(value, (dict, list)):
                # 如果值是字典或列表，则将其序列化为 JSON 字符串
                column_data.append(json.dumps(value, ensure_ascii=False))
            else:
                # 其他类型的值，确保转换为字符串
                column_data.append(str(value))
        df_data[lang] = column_data

    df = pd.DataFrame(df_data)

    # 4. 保存到文件
    try:
        if output_file.endswith('.xlsx'):
            # 对于 Excel 文件，需要特殊处理以在第一行添加注意事项
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill
            
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            
            # 在第一行第一列添加翻译注意事项
            notice_text = "请注意：文案中有 {}包裹的内容不能被翻译。比如： By continuing, you agree to our {userAgreement}, {privacyPolicy} and {communityGuidelines}. 其中{userAgreement} {privacyPolicy} {communityGuidelines} 是占位符，在其他语言下需要保持原样，不能被翻译。"
            
            # 先添加 DataFrame 数据（包括表头）从第二行开始
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置注意事项文本和样式
            ws['A1'] = notice_text
            
            # 设置第一行第一列的字体为红色和粗体，并添加浅灰色背景
            red_font = Font(color="FF0000", bold=True, size=11)
            light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            ws['A1'].font = red_font
            ws['A1'].fill = light_gray_fill
            
            # 调整列宽以适应长文本
            ws.column_dimensions['A'].width = 80
            
            # 保存工作簿
            wb.save(output_file)
        elif output_file.endswith('.csv'):
            # 对于 CSV 文件，先写入注意事项，再写入数据
            notice_text = "请注意：文案中有 {}包裹的内容不能被翻译。比如： By continuing, you agree to our {userAgreement}, {privacyPolicy} and {communityGuidelines}. 其中{userAgreement} {privacyPolicy} {communityGuidelines} 是占位符，在其他语言下需要保持原样，不能被翻译。"
            
            with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
                # 写入注意事项作为第一行
                f.write(f'"{notice_text}"\n')
                # 写入 DataFrame 数据
                df.to_csv(f, index=False, header=True)
        else:
            print(f"Error: Unsupported output file format. Please use .xlsx or .csv. Defaulting to .xlsx")
            # 默认使用 xlsx 格式
            output_file_xlsx = output_file + '.xlsx' if '.' not in output_file else output_file.split('.')[0] + '.xlsx'
            
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            
            notice_text = "请注意：文案中有 {}包裹的内容不能被翻译。比如： By continuing, you agree to our {userAgreement}, {privacyPolicy} and {communityGuidelines}. 其中{userAgreement} {privacyPolicy} {communityGuidelines} 是占位符，在其他语言下需要保持原样，不能被翻译。"
            
            # 先添加 DataFrame 数据（包括表头）从第二行开始
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置注意事项文本和样式
            ws['A1'] = notice_text
            
            # 设置第一行第一列的字体为红色和粗体，并添加浅灰色背景
            red_font = Font(color="FF0000", bold=True, size=11)
            light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            ws['A1'].font = red_font
            ws['A1'].fill = light_gray_fill
            
            # 调整列宽以适应长文本
            ws.column_dimensions['A'].width = 80
            
            wb.save(output_file_xlsx)
            
        print(f"Translations successfully exported to {output_file}")
    except Exception as e:
        print(f"Error writing to output file {output_file}: {e}")


def import_translations(input_file, translations_dir):
    print(f"Importing translations from {input_file} to {translations_dir}...")

    # 确保 translations_dir 存在，如果不存在则创建
    if not os.path.exists(translations_dir):
        try:
            os.makedirs(translations_dir)
            print(f"Created directory: {translations_dir}")
        except Exception as e:
            print(f"Error creating directory {translations_dir}: {e}")
            return

    # 1. 读取表格数据
    try:
        if input_file.endswith('.xlsx'):
            # 跳过第一行（注意事项），从第二行开始读取数据
            df = pd.read_excel(input_file, skiprows=1)
        elif input_file.endswith('.csv'):
            # 跳过第一行（注意事项），从第二行开始读取数据
            df = pd.read_csv(input_file, skiprows=1)
        else:
            print(f"Error: Unsupported input file format. Please use .xlsx or .csv.")
            return
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found.")
        return
    except Exception as e:
        print(f"Error reading input file {input_file}: {e}")
        return

    # 检查 'key' 列是否存在
    if 'key' not in df.columns:
        print(f"Error: 'key' column not found in {input_file}.")
        return

    # 2. 获取语言列 (除了 'key' 列之外的所有列)
    language_columns = [col for col in df.columns if col != 'key']

    if not language_columns:
        print(f"No language columns found in {input_file} (expected columns like 'en', 'zh', etc. besides 'key').")
        return

    # 3. 遍历每种语言并生成 .json 文件
    for lang_code in language_columns:
        translations_file_path = os.path.join(translations_dir, f'{lang_code}.json')
        # 尝试读取现有的 .json 文件以支持增量更新
        translations = {}
        if os.path.exists(translations_file_path):
            try:
                with open(translations_file_path, 'r', encoding='utf-8') as f_existing:
                    translations = json.load(f_existing)
            except Exception as e:
                print(
                    f"Warning: Could not read existing file {translations_file_path} for merging. It will be overwritten. Error: {e}")
                translations = {}  # 如果读取失败，则重置为空字典，相当于覆盖写入
        else:
            translations = {}  # 如果文件不存在，则初始化为空字典
        for index, row in df.iterrows():
            key = row['key']
            # 处理 NaN 或 None 值，转换为空字符串
            translation_value = row[lang_code]
            if pd.isna(translation_value):
                translation_value = ""
            else:
                # 尝试将字符串解析为 JSON 对象
                if isinstance(translation_value, str):
                    try:
                        # 只有当字符串看起来像一个JSON对象或数组时才尝试解析
                        if (translation_value.strip().startswith('{') and translation_value.strip().endswith('}')) or \
                                (translation_value.strip().startswith('[') and translation_value.strip().endswith(']')):
                            parsed_json = json.loads(translation_value)
                            translations[key] = parsed_json
                        else:
                            translations[key] = str(translation_value)
                    except json.JSONDecodeError:
                        # 如果解析失败，则保持为字符串
                        translations[key] = str(translation_value)
                else:
                    # 如果不是字符串（例如，已经是数字或布尔值），直接使用
                    translations[key] = translation_value
            # 对于表格中没有的 key，但存在于原 .json 文件中的，保留它们
            # (这一步通过先加载现有 translations 已经隐式完成)

        # 写入 .json 文件
        try:
            with open(translations_file_path, 'w', encoding='utf-8') as f:
                json.dump(translations, f, ensure_ascii=False, indent=2)
            print(f"Successfully imported translations for '{lang_code}' to {translations_file_path}")
        except Exception as e:
            print(f"Error writing to {translations_file_path}: {e}")

    print("Import process completed.")


def main():
    parser = argparse.ArgumentParser(description='Flutter Internationalization (i18n) Manager for .json files.')
    subparsers = parser.add_subparsers(dest='command', help='Available commands', required=True)

    # Export command
    parser_export = subparsers.add_parser('export', help='Export translations from .json files to a spreadsheet.')
    parser_export.add_argument('--translations_dir', type=str, default=DEFAULT_TRANSLATIONS_DIR,
                               help=f'Directory containing .json files (default: {DEFAULT_TRANSLATIONS_DIR})')
    parser_export.add_argument('--output', type=str, default='translations.xlsx',
                               help='Output spreadsheet file (e.g., translations.xlsx or translations.csv)')
    parser_export.set_defaults(func=lambda args: export_translations(args.translations_dir, args.output))

    # Import command
    parser_import = subparsers.add_parser('import', help='Import translations from a spreadsheet to .json files.')
    parser_import.add_argument('--translations_dir', type=str, default=DEFAULT_TRANSLATIONS_DIR,
                               help=f'Directory to save .json files (default: {DEFAULT_TRANSLATIONS_DIR})')
    parser_import.add_argument('--input', type=str, required=True,
                               help='Input spreadsheet file (e.g., translations.xlsx or translations.csv)')
    parser_import.set_defaults(func=lambda args: import_translations(args.input, args.translations_dir))

    args = parser.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()

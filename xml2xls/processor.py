import os
import re
import argparse
from openpyxl import Workbook, load_workbook

"""
Android字符串资源处理器

导出命令：
python3 processor.py --export res_dir translations.xlsx

导入命令：
python3 processor.py --import res_dir translations.xlsx --mode [full|partial]

参数说明：
  res_dir     资源目录路径（包含values/values-xx的文件夹）
  excel_file  Excel文件路径

选项：
  --export    导出到Excel（生成包含完整翻译和未翻译项的工作簿）
  --import    从Excel导入（支持两种模式）
  --mode      导入模式选择：
              full    - 从主表(All Translations)导入（默认）
              partial - 仅从未翻译表(Untranslated)导入

使用示例：
1. 导出所有翻译（含未翻译项）：
python3 processor.py --export app/src/main/res translations.xlsx

2. 导入完整翻译表（All Translations）数据：
python3 processor.py --import app/src/main/res translations.xlsx --mode full

3. 仅导入未翻译表（Untranslated）数据：
python3 processor.py --import app/src/main/res translations.xlsx --mode partial

"""


def export_to_excel(res_dir, output_file):
    """增强版导出功能，包含未翻译统计"""
    try:
        # 获取默认语言内容
        default_order, default_data = parse_strings_xml(os.path.join(res_dir, 'values', 'strings.xml'))
        all_langs = {'en': default_data}

        # 收集其他语言数据
        lang_codes = []
        for d in os.listdir(res_dir):
            if d.startswith('values-'):
                lang_code = d.replace('values-', '', 1)
                _, lang_data = parse_strings_xml(os.path.join(res_dir, d, 'strings.xml'))
                all_langs[lang_code] = lang_data
                lang_codes.append(lang_code)

        # 创建Excel文件
        wb = Workbook()

        # Sheet1: 完整翻译表
        main_sheet = wb.active
        main_sheet.title = "All Translations"
        headers = ['key', 'en'] + lang_codes
        main_sheet.append(headers)

        # 填充主表数据
        for key in default_order:
            row = [key, default_data.get(key, '')]
            for lc in lang_codes:
                row.append(all_langs[lc].get(key, ''))
            main_sheet.append(row)

        # Sheet2: 合并未翻译项（仅显示缺少翻译的条目）
        untrans_sheet = wb.create_sheet(title="Untranslated")
        untrans_sheet.append(headers)  # 保持相同表头

        # 构建未翻译数据（仅记录存在的翻译）
        for key in default_order:
            has_untranslated = False
            row_data = {'key': key, 'en': default_data.get(key, '')}

            # 检查每个语言的翻译状态
            for lc in lang_codes:
                translated_value = all_langs[lc].get(key, None)
                if translated_value is None:
                    has_untranslated = True
                    row_data[lc] = ''
                else:
                    row_data[lc] = translated_value

            # 仅添加有未翻译项的行
            if has_untranslated:
                row = [row_data['key'], row_data['en']]
                row += [row_data[lc] for lc in lang_codes]
                untrans_sheet.append(row)

        # 删除默认Sheet（如果存在）
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(output_file)
        print(f"导出成功：{output_file}")

    except Exception as e:
        print(f"导出失败：{str(e)}")
        raise


def import_from_excel(res_dir, input_file, mode='full'):
    """智能导入，支持选择数据源"""
    try:
        wb = load_workbook(input_file)
        lang_data = {}

        if mode == 'partial':
            if 'Untranslated' not in wb.sheetnames:
                raise ValueError("未找到未翻译工作表")

            sheet = wb['Untranslated']
            headers = [cell.value for cell in sheet[1]]
            all_lang_codes = headers[1:]  # 包含'en'

            # 处理所有语言列（包括en）
            for row in sheet.iter_rows(min_row=2):
                key = row[0].value
                if not key:
                    continue

                # 遍历所有语言列（从第2列开始）
                for idx, lang_code in enumerate(all_lang_codes, start=1):
                    cell = row[idx]
                    if cell.value:
                        lang_data.setdefault(lang_code, {})[key] = str(cell.value).strip()
        else:
            # 处理主表
            if 'All Translations' not in wb.sheetnames:
                raise ValueError("未找到主工作表")

            sheet = wb['All Translations']
            headers = [cell.value for cell in sheet[1]]
            lang_codes = headers[1:]  # 跳过key列

            for row in sheet.iter_rows(min_row=2):
                key = row[0].value
                if not key:
                    continue

                for idx, lc in enumerate(lang_codes, start=1):
                    value = row[idx].value
                    if value:
                        lang_data.setdefault(lc, {})[key] = str(value).strip()

        # 写入各语言文件
        for lang_code, data in lang_data.items():
            dir_name = 'values' if lang_code == 'en' else f'values-{lang_code}'
            xml_path = os.path.join(res_dir, dir_name, 'strings.xml')

            # 合并数据并写入
            existing_order, existing_data = parse_strings_xml(xml_path)
            merged_data = existing_data.copy()
            merged_data.update(data)

            write_strings_xml(xml_path, merged_data)

        print(f"导入成功！\n模式：{mode} \n更新语言：{list(lang_data.keys())}")
    except Exception as e:
        print(f"导入失败：{str(e)}")
        raise


def parse_strings_xml(xml_path):
    """解析XML获取原始内容（保留CDATA等特殊格式）"""
    order = []
    strings = {}
    if not os.path.exists(xml_path):
        return order, strings

    with open(xml_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 增强型正则表达式，匹配完整字符串定义
    pattern = re.compile(
        r'<!--(.*?)-->|'  # 匹配注释
        r'<string\s+name="([^"]+)"\s*>(.*?)</string>|'
        r'(</?resources>)',  # 匹配根标签
        re.DOTALL
    )

    for match in re.finditer(pattern, content):
        if match.group(1):  # 注释
            continue
        elif match.group(2):  # string标签
            name = match.group(2)
            value = match.group(3).strip()
            order.append(name)
            strings[name] = value
        elif match.group(4):  # resources标签
            continue  # 根标签不处理

    return order, strings


def write_strings_xml(xml_path, data):
    """智能合并写入XML（保留注释等其他内容）"""
    # 读取原始文件内容
    if os.path.exists(xml_path):
        with open(xml_path, 'r', encoding='utf-8') as f:
            content = f.read()
    else:
        content = '<?xml version="1.0" encoding="utf-8"?>\n<resources>\n</resources>'

    # 获取现有字符串和顺序
    existing_order, existing_data = parse_strings_xml(xml_path)

    # 合并数据：更新现有值，保留未修改内容
    merged_data = existing_data.copy()
    merged_data.update(data)
    new_order = existing_order + [k for k in data if k not in existing_order]

    # 构建替换字典（包含原始格式）
    replace_dict = {}
    for match in re.finditer(r'<string\s+name="([^"]+)"\s*>(.*?)</string>', content, re.DOTALL):
        name = match.group(1)
        if name in merged_data:
            old_content = match.group(0)
            new_content = f'<string name="{name}">{merged_data[name]}</string>'
            replace_dict[old_content] = new_content

    # 执行替换
    for old, new in replace_dict.items():
        content = content.replace(old, new, 1)

    # 添加新条目（追加到文件末尾前）
    new_entries = []
    for name in new_order:
        if name not in existing_data:
            new_entries.append(f'    <string name="{name}">{data[name]}</string>')

    if new_entries:
        # 在最后一个</string>后或</resources>前插入
        # 查找最后一个</string>的结束位置
        last_string_end = 0
        last_string_pos = content.rfind('</string>')
        if last_string_pos != -1:
            last_string_end = last_string_pos + len('</string>')

        # 查找</resources>的位置
        resources_end_pos = content.find('</resources>')

        # 确定插入位置（优先插入在最后一个</string>之后）
        insert_pos = last_string_end if last_string_end > 0 else resources_end_pos

        # 构建插入文本（包含换行格式）
        insert_text = '\n' + '\n'.join(f'{entry}' for entry in new_entries)

        # 执行插入（在插入位置后添加内容）
        content = content[:insert_pos] + insert_text + content[insert_pos:]

    # 处理xliff命名空间
    if 'xliff:' in content and 'xmlns:xliff' not in content:
        content = content.replace('<resources>',
                                  '<resources xmlns:xliff="urn:oasis:names:tc:xliff:document:1.2">', 1)

    # 写入文件（保留原始格式）
    os.makedirs(os.path.dirname(xml_path), exist_ok=True)
    with open(xml_path, 'w', encoding='utf-8') as f:
        f.write(content)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawTextHelpFormatter,
        description='''
Android字符串资源处理器
  导出/导入操作需配合 --export 或 --import 参数使用
  导出命令：
    python3 processor.py --export res_dir excel_file
  导入命令：
    python3 processor.py --import res_dir excel_file --mode [full|partial]
  ''',
        epilog='''
使用示例：
  1. 导出所有翻译（含未翻译项）：
     python3 processor.py --export app/src/main/res translations.xlsx

  2. 导入完整翻译表（All Translations）数据：
     python3 processor.py --import app/src/main/res translations.xlsx --mode full

  3. 仅导入未翻译表（Untranslated）数据：
     python3 processor.py --import app/src/main/res translations.xlsx --mode partial
        '''
    )

    parser.add_argument('--export', action='store_true',
                        help='导出到Excel（生成完整翻译表和未翻译项表）')
    parser.add_argument('--import', action='store_true', dest='import_',
                        help='从Excel导入（需配合 --mode 选择数据源, 支持两种模式）')
    parser.add_argument('--mode', choices=['full', 'partial'], default='full',
                        help='''
导入模式选择：                      
  full - 从主表(All Translations)导入（默认）
  partial - 仅从未翻译表(Untranslated)导入
    ''')
    parser.add_argument('res_dir', help='资源目录路径（包含 values/values-xx 的文件夹）')
    parser.add_argument('excel_file', help='Excel文件路径（输入/输出）')

    args = parser.parse_args()

    if args.export:
        export_to_excel(args.res_dir, args.excel_file)
    elif args.import_:
        import_from_excel(args.res_dir, args.excel_file, args.mode)
    else:
        print("请使用--export或--import参数")
